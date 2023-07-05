from openpyxl import load_workbook

def read_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    for d in ws.iter_rows(min_row=2,values_only=True):
        yield d

